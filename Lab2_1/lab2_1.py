from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import QVBoxLayout, QDialog, QWidget, QMessageBox, QTableWidgetItem, QCheckBox, QHBoxLayout
from UI import Ui_MainWindow
import openpyxl
import os

class CheckBoxWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.checkBoxLayout = QHBoxLayout(self)
        self.checkBoxLayout.setAlignment(QtCore.Qt.AlignCenter)
        self.checkBox = QCheckBox()
        self.checkBox.setChecked(False)
        self.checkBoxLayout.addWidget(self.checkBox)
        self.setLayout(self.checkBoxLayout)

class DialogChange(QDialog):
    def __init__(self, title, ObjectChange):
        super().__init__()
        self.setWindowIcon(QtGui.QIcon('./pictures/update.png'))
        self.title = title
        self.ObjectChange = ObjectChange
        self.newName = ''
        
        self.setFixedSize(250, 200)
        self.setWindowTitle(title)
        layout = QVBoxLayout(self)
        self.name_label = QtWidgets.QLabel("Enter new name:")
        self.name_edit = QtWidgets.QLineEdit()
        self.name_edit.setPlaceholderText(f"New name {self.ObjectChange}")
        self.button_1 = QtWidgets.QPushButton("Submit")
        self.button_2 = QtWidgets.QPushButton("Cancel")
        self.name_label.setStyleSheet("font: 12pt \"Segoe UI\";")
        self.name_edit.setStyleSheet("font: 12pt \"Segoe UI\";")
        self.button_1.setStyleSheet("font: 12pt \"Segoe UI\";")
        self.button_2.setStyleSheet("font: 12pt \"Segoe UI\";")
        layout.addWidget(self.name_label)
        layout.addWidget(self.name_edit)
        layout.addWidget(self.button_1)
        layout.addWidget(self.button_2)
        
        self.button_1.clicked.connect(self.getNewName)
        self.button_2.clicked.connect(self.closeDiaglog)
    
    def getNewName(self):
        self.newName = self.name_edit.text()
        if (self.newName != ''):
            self.accept()
        else:
            QMessageBox.critical(self, "Error!", "Name cannot be empty!")
            
    def closeDiaglog(self):
        self.newName = ''
        self.accept()
            
class MAIN_HANDLE(Ui_MainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.setupUi(MainWindow)
        MainWindow.setWindowIcon(QtGui.QIcon('./pictures/admin.png'))
        MainWindow.setWindowTitle("Administrator")
        self.rowHeight = 60
        self.columnWidth = 60
        self.tableFilePath = "./table.xlsx"
        self.listObject = []
        self.listSubject = []
        self.checkBoxBinVals  = []
        self.listCheckBox = []
        
        self.getTableFromFile()
        self.setTableAndComboBox()

        self.Add_O.clicked.connect(self.addNewObject)
        self.Add_S.clicked.connect(self.addNewSubject)
        self.Delete_S.clicked.connect(self.delSubject)
        self.Delete_O.clicked.connect(self.delObject)
        self.Change_S.clicked.connect(self.changeSubject)
        self.Change_O.clicked.connect(self.changeObject)
        self.Save.clicked.connect(self.saveTableToFile)
        self.Save.setDefault(True)

    def setTableAndComboBox(self):
        columns = len(self.listObject)
        rows = len(self.listSubject)
        
        self.tableWidget.setRowCount(rows)
        self.tableWidget.setColumnCount(columns)
        self.tableWidget.setHorizontalHeaderLabels(self.listObject)
        self.tableWidget.setVerticalHeaderLabels(self.listSubject)

        for row in range(rows):
            row_check_box = []
            self.tableWidget.setRowHeight(row, self.rowHeight)
            for column in range(columns):
                self.tableWidget.setColumnWidth(column, self.columnWidth)               
                check_box = CheckBoxWidget()
                check_box.checkBox.setChecked(self.checkBoxBinVals[row][column])
                self.tableWidget.setCellWidget(row, column, check_box)
                row_check_box.append(check_box)
                
            self.listCheckBox.append(row_check_box)

        for i in range(len(self.listSubject)):
            self.comboBox.addItem(self.listSubject[i])
        
        for i in range(len(self.listObject)):
            self.comboBox_2.addItem(self.listObject[i])
    
    def getTableFromFile(self):
        if (os.path.isfile(self.tableFilePath)):
            wb = openpyxl.load_workbook(self.tableFilePath)
            sheet = wb.active
            for subjectNum in range (2, sheet.max_row+1):
                cell = sheet.cell(subjectNum, 1)
                self.listSubject.append(cell.value)
            for objectNum in range (2, sheet.max_column+1):
                cell = sheet.cell(1, objectNum)
                self.listObject.append(cell.value)
            for subjectNum in range (2, sheet.max_row+1):
                rowBinVals = []
                for objectNum in range (2, sheet.max_column+1):
                    cell = sheet.cell(subjectNum, objectNum)
                    rowBinVals.append(cell.value)
                self.checkBoxBinVals.append(rowBinVals)
            wb.close()
        else:
            wb = openpyxl.Workbook()
            wb.save(self.tableFilePath)
            self.ShowStatus("Information!","file created no info in file yet!", True)
            wb.close()

    def saveTableToFile(self):
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            rows = len(self.listSubject)
            columns = len(self.listObject)
            self.getValueCheckBox()

            for subjectNum in range (0, rows):
                cell = sheet.cell(subjectNum+2, 1)
                cell.value = self.listSubject[subjectNum]

            for objectNum in range (0, columns):
                cell = sheet.cell(1, objectNum+2)
                cell.value = self.listObject[objectNum]

            for subjectNum in range (0, rows):
                for objectNum in range (0, columns):
                    cell = sheet.cell(subjectNum+2, objectNum+2)
                    cell.value = self.checkBoxBinVals[subjectNum][objectNum]
            wb.save(self.tableFilePath)
            self.ShowStatus("Success!", "File saved successfully", True)
        except:
            self.ShowStatus("Error!", "File saved failed", False)
        finally:
            wb.close()
          
    def getValueCheckBox(self):
        self.checkBoxBinVals = []
        rows = len(self.listSubject)
        columns = len(self.listObject)
       
        for row in range(rows):
            columnVals = []
            for column in range(columns): 
                checkBox = self.tableWidget.cellWidget(row, column).layout().itemAt(0).widget()
                columnVals.append(int(checkBox.isChecked()))
            self.checkBoxBinVals.append(columnVals)
        
    def addNewSubject(self):
        nameSubject = self.subject.text()       
        if nameSubject and all(x.isalnum() for x in nameSubject):
            if nameSubject not in self.listSubject:
                row_position = self.tableWidget.rowCount()
                self.tableWidget.insertRow(row_position)
                columns = self.tableWidget.columnCount()
                self.tableWidget.setRowHeight(row_position, self.rowHeight)
                new_row_check_box = []
                for column in range(columns):
                    check_box = CheckBoxWidget()
                    self.tableWidget.setCellWidget(row_position, column, check_box)
                    new_row_check_box.append(check_box)
                self.listCheckBox.append(new_row_check_box)
                self.tableWidget.setVerticalHeaderItem(row_position, QTableWidgetItem(nameSubject))
                
                self.listSubject.append(nameSubject)
                self.comboBox.addItem(nameSubject)
                self.ShowStatus("Success!", "Added new object!", True)
            else:
                self.ShowStatus("Error!", "Subject already exists!", False)
        else:
            self.ShowStatus("Error!", "Invalid subject name! (can only contain letters or numbers)!", False)
        self.subject.setText("")

    def addNewObject(self):
        nameObject = self.object.text()
        if nameObject and len(nameObject) == 1 and nameObject.isalnum():
            if nameObject not in self.listObject:
                column_position = self.tableWidget.columnCount()
                self.tableWidget.insertColumn(column_position)
                rows = self.tableWidget.rowCount()
                self.tableWidget.setColumnWidth(column_position, self.columnWidth)
                for row in range(rows):
                    check_box = CheckBoxWidget()
                    self.tableWidget.setCellWidget(row, column_position, check_box)
                    self.listCheckBox[row].append(check_box)
                self.tableWidget.setHorizontalHeaderItem(column_position, QTableWidgetItem(nameObject))
                
                self.listObject.append(nameObject)
                self.comboBox_2.addItem(nameObject)
                self.ShowStatus("Success!", "Added new subject!", True)
            else:
                self.ShowStatus("Error!", "Object already exists!", False)
        else:
            self.ShowStatus("Error!", "Invalid object name! (can only contain letter or number)!", False)
        self.object.setText("")

    def delSubject(self):
        nameSubject = self.comboBox.currentText()
        IdSubject = self.comboBox.currentIndex()
        if nameSubject in self.listSubject and IdSubject >= 0:
            self.comboBox.removeItem(IdSubject)
            self.tableWidget.removeRow(IdSubject)
            self.listSubject.remove(nameSubject)
            self.listCheckBox.remove(self.listCheckBox[IdSubject])
            self.ShowStatus("Success!", "Deleted subject", True)
   
    def delObject(self):
        nameObject = self.comboBox_2.currentText()
        IdObject = self.comboBox_2.currentIndex()
        if nameObject in self.listObject and IdObject >= 0:
            self.comboBox_2.removeItem(IdObject)
            self.tableWidget.removeColumn(IdObject)
            self.listObject.remove(nameObject)
            for i in range(len(self.listCheckBox)):
                self.listCheckBox[i].remove(self.listCheckBox[i][IdObject])
            self.ShowStatus("Success!", "Deleted object", True)

    def changeSubject(self):
        IdSubject = self.comboBox.currentIndex()
        dialog = DialogChange("Edit subject", "subject")
        dialog.exec_()
        newName = dialog.newName

        if newName and all(x.isalnum() for x in newName):
            if (newName not in self.listSubject):
                self.comboBox.setItemText(IdSubject, newName)
                self.tableWidget.verticalHeaderItem(IdSubject).setText(newName)
                self.listSubject[IdSubject] = newName
                self.ShowStatus("Success!", "Updated new name Object!", True)
            else:
                self.ShowStatus("Error", "Update failed, new name already exists", False)
        elif newName:
            self.ShowStatus("Error!", "Invalid subject name! (can only contain letters or numbers)!", False)
            
    def changeObject(self):
        IdObject = self.comboBox_2.currentIndex()
        dialog = DialogChange("Edit object", "object")
        dialog.exec_()
        newName = dialog.newName
        
        if newName and len(newName) == 1 and newName.isalnum():
            if (newName and newName not in self.listObject):
                self.comboBox_2.setItemText(IdObject, newName)
                self.tableWidget.horizontalHeaderItem(IdObject).setText(newName)
                self.listObject[IdObject] = newName
                self.ShowStatus("Success!", "Updated new name Object!", True)
            else:
                self.ShowStatus("Error", "Update failed, new name already exists", False)
        elif newName:
            self.ShowStatus("Error!", "Invalid object name! (can only contain letter or number)!", False)
    
    def ShowStatus(self, Title, Message, Status):
        error = QMessageBox()
        error.setWindowTitle(Title)
        error.setText(Message)
        if not Status:
            error.setIcon(QMessageBox.Critical) 
        else:
            error.setIcon(QMessageBox.Information) 
        error.setStandardButtons(QMessageBox.Ok) 
        error.setDefaultButton(QMessageBox.Ok) 
        error.setStyleSheet("font: 14pt \"MV Boli\";\n""color: rgb(255, 0, 0);")
        error.setWindowIcon(QtGui.QIcon('./pictures/notifi.png'))
        error.exec_()
        

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = MAIN_HANDLE()
    MainWindow.show()
    sys.exit(app.exec_())